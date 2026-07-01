from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = "1F4E78"
INPUT_FILL = "FFF2CC"
NOTE_FILL = "F2F2F2"
INFO_FILL = "D9EAF7"
OUTPUT_FILL = "E2F0D9"
WHITE = "FFFFFF"
BLUE = "0000FF"
BLACK = "000000"
GRID = "BFBFBF"
INPUT_RESERVE_ROWS = 500


SHEET_NOTES = {
    "总体参数": "填写全局计算参数。时间单位为 s，长度单位为 m，质量单位为 kg。",
    "级事件": "填写各子级或助推器的点火、关机和分离时间。按 QJ 标准，分离规则为 t > ts 时剔除该级/助推器。",
    "不变质量": "填写结构、设备、有效载荷等不含推进剂和增压气体的部组件质量特性。",
    "可抛质量": "填写整流罩、热级环等按时间抛离的部件。无可抛部件时可留空。",
    "贮箱基本参数": "填写每个贮箱的基本参数。等效圆柱只需填写等效半径R和等效高度H；QJ第一类需补充几何表。",
    "贮箱几何_第一类": "按 QJ 1080A 图3填写第一类外轮廓半径模型参数。等效圆柱贮箱不需要填写本表。",
    "贮箱几何_第二类": "按 QJ 1080A 图4填写内部扣除体积半径模型参数。没有扣除体积时选择 否。",
    "外行导管": "填写外行导管内推进剂对 Jx 的贡献。没有外行导管时可留空或填0。",
    "运输状态": "填写运输状态部段组件质量特性，程序按部段编号合成运输状态结果。",
}


HEADER_COMMENTS = {
    "参数": "总体参数名称。程序按文字名称读取，如 飞行开始时间、飞行结束时间、时间步长、液位积分点数。",
    "值": "参数取值。请按本行单位填写；空值通常会被程序按默认值处理，但关键时间和密度不建议留空。",
    "单位": "变量单位说明。程序不做单位换算，所有输入应统一使用 kg、m、s、kg/m^3、kg·m^2。",
    "说明": "填写备注或工程来源说明，不参与计算。",
    "备注": "自由备注列，不参与计算。可记录数据来源、假设或校核说明。",
    "类型": "对象类型。选择 子级 或 助推器；用于生成贮箱ID、匹配级事件，并决定助推器 Y/Z 坐标是否叠加安装偏置。",
    "编号": "子级或助推器编号。用于和 不变质量、贮箱基本参数、级事件 等表按 类型+编号 匹配。",
    "所属编号": "该组件或贮箱所属的子级/助推器编号。与 类型、贮箱编号 一起生成贮箱ID：类型-所属编号-贮箱编号。",
    "贮箱编号": "同一子级/助推器内的贮箱序号。必须与几何表、外行导管表中的贮箱ID保持一致。",
    "名称": "对象名称，仅用于报表识别和关键事件说明，不参与数值计算。",
    "组件名称": "不变质量组件名称，仅用于识别数据来源，不参与数值计算。",
    "几何类型": "贮箱几何计算方式。等效圆柱：使用等效半径R和等效高度H；QJ第一类：使用贮箱几何_第一类，必要时叠加第二类扣除。",
    "等效半径R": "等效圆柱半径，单位 m。仅几何类型为 等效圆柱 时使用，截面积 A=pi*R^2。",
    "等效高度H": "等效圆柱高度，单位 m。仅几何类型为 等效圆柱 时使用，液位从 0 积分到该高度。",
    "M": "质量，单位 kg。用于不变质量、可抛质量或运输状态组件。应为该组件当前输入状态下的实际质量。",
    "X": "质心 X 坐标，单位 m。不变质量/可抛质量使用全局 OXYZ 坐标系，O 为火箭理论顶点，OX 指向火箭底部为正。",
    "Y": "质心 Y 坐标，单位 m。不变质量/可抛质量使用全局 OXYZ 坐标系。",
    "Z": "质心 Z 坐标，单位 m。不变质量/可抛质量使用全局 OXYZ 坐标系。",
    "Jx": "绕组件自身质心主轴 X 的中心主转动惯量，单位 kg·m^2。程序合成总惯量时会自动应用平行轴定理。",
    "Jy": "绕组件自身质心主轴 Y 的中心主转动惯量，单位 kg·m^2。程序合成总惯量时会自动应用平行轴定理。",
    "Jz": "绕组件自身质心主轴 Z 的中心主转动惯量，单位 kg·m^2。程序假设惯性积为 0。",
    "点火时间tf": "发动机开始工作时间，单位 s。贮箱可变质量在 t<tf 使用加注状态 Mf/Xf/Jyf；tf<=t<=tb 进入工作段消耗。",
    "关机时间tb": "发动机关机时间，单位 s。要求 tf<=tb<=ts；t>tb 且 t<=ts 时贮箱使用关机剩余状态 Mr/Xr/Jyr。",
    "分离时间ts": "该级/助推器分离时间，单位 s。按 QJ 标准，程序采用 t<=ts 保留，t>ts 剔除。",
    "抛离时间": "可抛部件抛离时间，单位 s。按 QJ 标准，t<=抛离时间 时保留，t>抛离时间 时剔除。",
    "推进剂密度rho_phi": "推进剂密度 rho_phi，单位 kg/m^3。用于 M_phi=rho_phi*V_phi，也用于气体惯量密度比例换算。必须大于0。",
    "气体密度rho_g": "增压气体密度 rho_g，单位 kg/m^3。用于 M_g=rho_g*V_g；若忽略气体质量可填0。",
    "秒流量mdot": "推进剂/可变质量工作段消耗率，单位 kg/s。程序在 tf<=t<=tb 使用 Mef - mdot*(t-tf) 计算目标质量。",
    "加注量Mf": "QJ 中 Mf。发动机点火前贮箱内可变质量，单位 kg。t<tf 时使用该质量，并优先使用 Xf/Jyf；若 Xf/Jyf 留空则按 Mf 查液位表。",
    "关机剩余量Mr": "QJ 中 Mr。发动机关机后剩余可变质量，单位 kg。t>tb 且 t<=ts 时使用该质量，并优先使用 Xr/Jyr。",
    "启动后剩余量Mef": "QJ 式(92)中的 Mef。启动活门打开后进入正常工作段消耗计算时的剩余可变质量，单位 kg。若留空，程序默认 Mef=Mf。",
    "加注后Xf": "QJ 中 Xf。点火前加注状态可变质量在贮箱局部坐标系下的 X 坐标，单位 m；全局 X=Loki-Xf。留空则按 Mf 查液位表。",
    "关机后Xr": "QJ 中 Xr。关机后剩余状态可变质量在贮箱局部坐标系下的 X 坐标，单位 m；全局 X=Loki-Xr。留空则按 Mr 查液位表。",
    "加注后Jyf": "QJ 中 Jyf。点火前加注状态可变质量绕贮箱局部 Y 轴的中心主惯量，单位 kg·m^2；Jz 默认等于 Jy。留空则按 Mf 查液位表。",
    "关机后Jyr": "QJ 中 Jyr。关机后剩余状态可变质量绕贮箱局部 Y 轴的中心主惯量，单位 kg·m^2；Jz 默认等于 Jy。留空则按 Mr 查液位表。",
    "坐标原点X或Loki": "QJ 中 Loki/Lx0ji。贮箱局部坐标原点到全局 O 点的 X 向距离，单位 m。程序使用 X_global=abs(Loki)-X_local。",
    "坐标原点Y": "助推器局部坐标原点在全局 OXYZ 中的 Y 坐标，单位 m。仅 类型=助推器 时叠加到贮箱可变质量 Y 坐标。",
    "坐标原点Z": "助推器局部坐标原点在全局 OXYZ 中的 Z 坐标，单位 m。仅 类型=助推器 时叠加到贮箱可变质量 Z 坐标。",
    "箱底hb": "QJ 5.2.5 中 h_bki。箱底平面在贮箱局部 X 轴上的坐标，单位 m。用于输出 从箱底算起 的液位结果。",
    "总高H": "贮箱总高备用字段，单位 m。当 QJ 几何高度或等效高度无效时用于兜底确定积分高度。",
    "总容积Vo": "QJ 中 Voki。用于 V_g=Vo-V_phi 计算增压气体体积。标准中 Vo 应为式(54)积分到满箱高度得到的总容积；留空则使用液位表最大 V_phi。",
    "满箱Xo": "QJ 中 Xoki。贮箱注满推进剂后总体积质心在贮箱局部 X 轴上的坐标，单位 m。用于反推气体质心 X_g。留空则使用最大液位 X_phi。",
    "满箱Jyo": "QJ 中 Jyoki。贮箱注满推进剂后绕局部 Y 轴的中心主惯量，单位 kg·m^2。用于估算增压气体 Jy/Jz。留空则使用最大液位 Jy_phi。",
    "是否启用扣除": "第二类几何开关。是：R2(x) 作为内部扣除半径参与 A=pi*(R1^2-R2^2)；否：R2(x)=0。",
    "semi_a11": "第一类几何图3中的椭圆横向半轴 a11，单位 m。用于 ellipse14/ellipse16 段，程序用 semi_a11/b 形成比例系数。",
    "semi_a22": "第二类几何图4中的椭圆横向半轴 a22，单位 m。用于第二类椭圆扣除段，程序用 semi_a22/b 形成比例系数。",
    "b11": "第一类几何椭圆段参数 b11，单位 m。用于 H13~H14 段的椭圆半径计算。",
    "b12": "第一类几何椭圆段参数 b12，单位 m。用于 H15~H16 段的椭圆半径计算。",
    "b21": "第二类几何椭圆扣除段参数 b21，单位 m。用于 H21~H22 段。",
    "b22": "第二类几何椭圆扣除段参数 b22，单位 m。用于 H27~H28 段。",
    "h11": "第一类几何圆弧段圆心相关 X 坐标，单位 m。用于 arc12 半径公式。",
    "h12": "第一类几何圆弧段圆心相关 X 坐标，单位 m。用于 arc18 半径公式。",
    "h21": "第二类几何圆弧扣除段圆心相关 X 坐标，单位 m。用于 arc24 半径公式。",
    "h22": "第二类几何圆弧扣除段圆心相关 X 坐标，单位 m。用于 arc26 半径公式。",
    "beta11_deg": "第一类几何锥段角 beta11，单位 deg。程序内部使用 tan(beta11) 作为 H12~H13 段半径变化斜率。",
    "beta12_deg": "第一类几何锥段角 beta12，单位 deg。程序内部使用 tan(beta12) 作为 H16~H17 段半径变化斜率。",
    "beta21_deg": "第二类几何锥段角 beta21，单位 deg。程序内部使用 tan(beta21) 作为 H22~H23 扣除半径变化斜率。",
    "beta22_deg": "第二类几何锥段角 beta22，单位 deg。程序内部使用 tan(beta22) 作为 H26~H27 扣除半径变化斜率。",
    "导管编号": "外行导管序号，仅用于识别。同一贮箱可填写多条导管，程序会合成导管推进剂的 Y/Z/Jx 贡献。",
    "m_phi": "外行导管内推进剂质量，单位 kg。QJ 假设主体推进剂绕 X 轴惯量不计，Jx 主要由外行导管推进剂贡献。",
    "Y_phi": "外行导管内推进剂质心在贮箱局部 Y 轴上的坐标，单位 m。用于计算可变质量 Y 偏心。",
    "Z_phi": "外行导管内推进剂质心在贮箱局部 Z 轴上的坐标，单位 m。用于计算可变质量 Z 偏心。",
    "Jx_phi": "外行导管内推进剂绕局部 X 轴的中心主惯量，单位 kg·m^2。程序将其作为贮箱可变质量 Jx 的主要贡献。",
    "部段编号": "运输状态部段编号。程序按该字段分组，合成每个部段的运输状态质量、质心和惯量。",
    "部段名称": "运输状态部段名称，仅用于输出识别。",
    "Xq": "部段坐标系 O_cq 原点在全局 OXYZ 中的 X 坐标，单位 m。当前运输状态结果输出保留该字段用于坐标转换参考。",
    "Xcq": "运输状态组件质心在部段局部 O_cqX_cq 轴上的坐标，单位 m。用于部段内质量特性合成。",
    "Ycq": "运输状态组件质心在部段局部 O_cqY_cq 轴上的坐标，单位 m。用于部段内质量特性合成。",
    "Zcq": "运输状态组件质心在部段局部 O_cqZ_cq 轴上的坐标，单位 m。用于部段内质量特性合成。",
    "Jxcq": "运输状态组件绕自身质心 X_cq 轴的中心主惯量，单位 kg·m^2。",
    "Jycq": "运输状态组件绕自身质心 Y_cq 轴的中心主惯量，单位 kg·m^2。",
    "Jzcq": "运输状态组件绕自身质心 Z_cq 轴的中心主惯量，单位 kg·m^2。",
}

for i in range(11, 17):
    HEADER_COMMENTS[f"R{i}"] = f"第一类几何外轮廓半径参数 R{i}，单位 m。按 QJ 图3分段公式参与 R1(x) 计算。"

for i in range(11, 20):
    HEADER_COMMENTS[f"H{i}"] = f"第一类几何外轮廓分段高度 H{i}，单位 m。定义 R1(x) 对应区段的上下边界。"

for i in range(21, 28):
    HEADER_COMMENTS[f"R{i}"] = f"第二类几何内部扣除半径参数 R{i}，单位 m。启用扣除时按 QJ 图4参与 R2(x) 计算。"

for i in range(21, 30):
    HEADER_COMMENTS[f"H{i}"] = f"第二类几何内部扣除分段高度 H{i}，单位 m。定义 R2(x) 对应区段的上下边界。"


def add_instructions_sheet(wb) -> None:
    if "填写说明" in wb.sheetnames:
        del wb["填写说明"]
    ws = wb.create_sheet("填写说明", 0)
    rows = [
        ["主题", "说明"],
        ["模板定位", "Excel 只负责输入和结果查看，复杂计算由 Python 完成。"],
        ["坐标系", "O 为火箭理论顶点；X 轴沿火箭中心线指向底部为正；惯性积按 0 处理。"],
        ["单位", "质量 kg；长度/坐标 m；时间 s；密度 kg/m^3；流量 kg/s；转动惯量 kg·m^2。"],
        ["填写顺序", "总体参数 -> 级事件 -> 不变质量 -> 贮箱基本参数 -> 贮箱几何/外行导管/可抛质量。"],
        ["贮箱ID", "规则为 类型-所属编号-贮箱编号，例如 子级-1-1、助推器-1-2。"],
        ["几何类型", "等效圆柱：填写等效半径R和等效高度H；QJ第一类：填写贮箱几何_第一类，必要时填写第二类扣除。"],
        ["级间分离", "按 QJ 标准，程序采用 t <= ts 保留，t > ts 剔除该级/助推器。"],
        ["可变质量分段", "t<tf 使用 Mf/Xf/Jyf；tf<=t<=tb 使用 Mef-mdot*(t-tf) 查液位表；tb<t<=ts 使用 Mr/Xr/Jyr；t>ts 为0。"],
        ["留空默认", "Mef 留空时默认等于 Mf；Xf/Xr/Jyf/Jyr 留空时分别按 Mf/Mr 在液位表中查最近状态；Vo/Xo/Jyo 留空时使用满液位积分结果。"],
        ["Vo关系", "QJ 标准中 Vo 为式(54)积分得到的总容积；程序允许手填 Vo 覆盖，用于气体体积 Vg=Vo-Vphi，并会提示与几何满箱体积偏差。"],
        ["外行导管", "按 QJ 假设，主体推进剂绕 X 轴惯量不计，Jx 主要来自外行导管推进剂；无外行导管可留空或填0。"],
        ["标准符号", "Mf=加注量；Mef=启动后剩余量；Mr=关机剩余量；Loki=贮箱局部原点到全局O点X向距离；Vo/Xo/Jyo=满箱参考体积、质心、惯量。"],
        ["局部到全局", "贮箱液位积分在局部 X 轴上进行；子级/芯级全局 X=abs(Loki)-X_local；助推器还会叠加坐标原点Y/Z。"],
        ["箱底液位", "箱底hb 对应 QJ h_bki，用于输出从箱底算起的 V_b、M_b、X_b、Jy_b、T_b，不改变主液位表积分。"],
        ["颜色", "浅黄色/蓝字为输入区；灰色为备注说明；深蓝色为表头。"],
        ["Starship样例", "input_starship_example.xlsx 是公开数据近似验证样例，不代表 SpaceX 官方内部质量特性模型。"],
    ]
    for r, row in enumerate(rows, 1):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    apply_grid(ws)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        row[0].fill = PatternFill("solid", fgColor=INFO_FILL)
        row[0].font = Font(name="Arial", bold=True, color=BLACK)
        row[1].alignment = Alignment(vertical="center", wrap_text=True)


WIDTHS = {
    "参数": 24,
    "值": 40,
    "单位": 12,
    "说明": 56,
    "备注": 46,
    "名称": 26,
    "组件名称": 28,
    "部段名称": 24,
    "贮箱ID": 18,
    "类型": 12,
    "几何类型": 16,
    "坐标原点X或Loki": 20,
    "推进剂密度rho_phi": 20,
    "气体密度rho_g": 18,
    "秒流量mdot": 16,
    "启动后剩余量Mef": 18,
    "加注后Xf": 16,
    "关机后Xr": 16,
    "加注后Jyf": 16,
    "关机后Jyr": 16,
    "等效半径R": 16,
    "等效高度H": 16,
    "总容积Vo": 16,
    "满箱Xo": 16,
    "满箱Jyo": 18,
    "事件类型": 18,
    "对象": 18,
    "事件时间": 16,
    "匹配时间": 16,
}


def apply_grid(ws, max_row: int | None = None, max_col: int | None = None) -> None:
    thin = Side(style="thin", color=GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def add_dropdown(ws, cell_range: str, options: list[str]) -> None:
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    dv.error = "请从下拉列表中选择。"
    dv.errorTitle = "无效选项"
    dv.prompt = "请选择：" + "、".join(options)
    dv.promptTitle = "可选值"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_decimal_validation(ws, cell_range: str, min_value: float = 0.0, allow_blank: bool = True) -> None:
    dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=str(min_value), allow_blank=allow_blank)
    dv.error = f"请输入大于等于 {min_value} 的数值。"
    dv.errorTitle = "数值范围错误"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_integer_validation(ws, cell_range: str, min_value: int = 0, allow_blank: bool = True) -> None:
    dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1=str(min_value), allow_blank=allow_blank)
    dv.error = f"请输入大于等于 {min_value} 的整数。"
    dv.errorTitle = "整数范围错误"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def header_map(ws) -> dict[str, int]:
    return {str(cell.value): cell.column for cell in ws[1] if cell.value is not None}


def range_for_col(ws, col_name: str, start: int = 2, end: int = 500) -> str | None:
    h = header_map(ws)
    if col_name not in h:
        return None
    letter = ws.cell(1, h[col_name]).column_letter
    return f"{letter}{start}:{letter}{end}"


def apply_validations(ws) -> None:
    for col in ["类型"]:
        rng = range_for_col(ws, col)
        if rng:
            add_dropdown(ws, rng, ["子级", "助推器"])
    rng = range_for_col(ws, "几何类型")
    if rng:
        add_dropdown(ws, rng, ["等效圆柱", "QJ第一类"])
    rng = range_for_col(ws, "是否启用扣除")
    if rng:
        add_dropdown(ws, rng, ["否", "是"])

    nonnegative_keywords = ["M", "质量", "密度", "秒流量", "加注量", "剩余量", "时间", "半径", "高度", "容积", "Jx", "Jy", "Jz", "R", "H", "h", "b", "semi_a"]
    integer_cols = {"编号", "所属编号", "贮箱编号", "导管编号"}
    for name, col in header_map(ws).items():
        rng = f"{ws.cell(1, col).column_letter}2:{ws.cell(1, col).column_letter}500"
        if name in integer_cols:
            add_integer_validation(ws, rng, 0)
        elif any(k in name for k in nonnegative_keywords):
            if name not in {"X", "Y", "Z", "Xcq", "Ycq", "Zcq", "坐标原点X或Loki", "坐标原点Y", "坐标原点Z"}:
                add_decimal_validation(ws, rng, 0)


def format_worksheet(ws) -> None:
    reserve_rows = max(INPUT_RESERVE_ROWS, ws.max_row)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A1:{ws.cell(1, ws.max_column).column_letter}{reserve_rows}"
    apply_grid(ws, reserve_rows, ws.max_column)
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = 34
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell.value in HEADER_COMMENTS:
            cell.comment = Comment(HEADER_COMMENTS[cell.value], "OpenCode")
    for row in ws.iter_rows(min_row=2, max_row=reserve_rows, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=BLUE)
            cell.fill = PatternFill("solid", fgColor=INPUT_FILL)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    if ws.max_column >= 1:
        for cell in ws.iter_cols(min_col=ws.max_column, max_col=ws.max_column, min_row=2, max_row=reserve_rows):
            for c in cell:
                c.fill = PatternFill("solid", fgColor=NOTE_FILL)
                c.font = Font(name="Arial", size=10, color=BLACK)
    for col in ws.columns:
        header = str(col[0].value) if col[0].value is not None else ""
        width = WIDTHS.get(header, max(14, min(max(len(str(c.value)) if c.value is not None else 0 for c in col[: min(len(col), 50)]) + 4, 42)))
        ws.column_dimensions[col[0].column_letter].width = width
    apply_validations(ws)


def add_sheet_notes(wb) -> None:
    for ws in wb.worksheets:
        if ws.title == "填写说明":
            continue
        note = SHEET_NOTES.get(ws.title)
        if note:
            ws["A1"].comment = Comment((ws["A1"].comment.text + "\n\n" if ws["A1"].comment else "") + note, "OpenCode")


def format_input_workbook(path: str) -> None:
    wb = load_workbook(path)
    add_instructions_sheet(wb)
    for ws in wb.worksheets:
        if ws.title != "填写说明":
            format_worksheet(ws)
    add_sheet_notes(wb)
    wb.save(path)


def format_output_workbook(path: str) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        max_row = max(ws.max_row + 20, 100)
        max_col = ws.max_column
        ws.auto_filter.ref = f"A1:{ws.cell(1, max_col).column_letter}{ws.max_row}"
        apply_grid(ws, max_row, max_col)
        ws.row_dimensions[1].height = 32
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.font = Font(name="Arial", size=10, color=BLACK)
                cell.fill = PatternFill("solid", fgColor=OUTPUT_FILL)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                header = str(ws.cell(1, cell.column).value)
                if header in {"M", "M_v", "M_phi", "M_g", "Jx", "Jy", "Jz", "Jx_v", "Jy_v", "Jz_v", "Jx_phi", "Jy_phi", "Jz_phi"}:
                    cell.number_format = '#,##0.000'
                elif header in {"X", "Y", "Z", "h", "X_v", "Y_v", "Z_v", "X_phi", "Y_phi", "Z_phi", "X_g"}:
                    cell.number_format = '0.000'
                elif header in {"t", "事件时间", "匹配时间", "前一时间", "后一时间"}:
                    cell.number_format = '0.0'
        for col in ws.columns:
            header = str(col[0].value) if col[0].value is not None else ""
            width = WIDTHS.get(header, max(14, min(max(len(str(c.value)) if c.value is not None else 0 for c in col[: min(len(col), 80)]) + 4, 42)))
            ws.column_dimensions[col[0].column_letter].width = width
    wb.save(path)
