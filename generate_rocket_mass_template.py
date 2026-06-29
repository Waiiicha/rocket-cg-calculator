from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = r"E:\rocket_program\cg_inertia_calculator\液体火箭质量质心转动惯量计算模板.xlsx"

BLUE = "0000FF"
BLACK = "000000"
GREEN = "008000"
WHITE = "FFFFFF"
HEADER = "1F4E78"
SUBHEADER = "D9EAF7"
INPUT_FILL = "FFF2CC"
CALC_FILL = "E2F0D9"
NOTE_FILL = "F2F2F2"
GRID = "D9D9D9"


def style_sheet(ws, widths=None):
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows():
        for c in row:
            c.font = Font(name="Arial", size=10, color=BLACK)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.font = Font(name="Arial", size=14, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for c in ws[3]:
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width


def title(ws, text, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1, text)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[3].height = 36


def mark_inputs(ws, min_row, max_row, cols):
    for row in range(min_row, max_row + 1):
        for col in cols:
            c = ws.cell(row, col)
            c.font = Font(name="Arial", size=10, color=BLUE)
            c.fill = PatternFill("solid", fgColor=INPUT_FILL)


def set_formula(cell, formula, link=False):
    cell.value = formula
    cell.font = Font(name="Arial", size=10, color=GREEN if link else BLACK)
    cell.fill = PatternFill("solid", fgColor=CALC_FILL)


def build():
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("说明")
    title(ws, "液体火箭质量、质心和转动惯量计算模板", 6)
    rows = [
        ["标准依据", "QJ 1080A-1997 液体火箭质量、质心和转动惯量计算方法", "", "", "", ""],
        ["坐标系", "OXYZ，原点为火箭理论顶点，OX 指向火箭底部为正；惯性积按 0 处理。", "", "", "", ""],
        ["颜色约定", "蓝色/浅黄为输入；黑色/浅绿为公式；绿色字体为跨表引用。", "", "", "", ""],
        ["使用流程", "先填总体参数、不变质量、贮箱参数、可抛质量；再查看计算与输出页。", "", "", "", ""],
        ["贮箱积分", "模板预留 101 个液位离散点，采用梯形积分近似连续积分。", "", "", "", ""],
        ["注意", "本模板不考虑火箭机动过载影响；复杂非轴对称结构需单独校核。", "", "", "", ""],
    ]
    for r, row in enumerate(rows, 3):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    style_sheet(ws, {"A": 18, "B": 90, "C": 15, "D": 15, "E": 15, "F": 15})

    ws = wb.create_sheet("输入_总体参数")
    title(ws, "输入_总体参数", 10)
    headers = ["参数", "值", "单位", "说明", "", "级/助推器", "点火 tf", "关机 tb", "分离 ts", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    data = [
        ["火箭总级数 Nr", 3, "-", "通用模板默认 3，可修改"],
        ["助推器数量 Nzt", 0, "-", "无助推器填 0"],
        ["液位离散点数", 101, "-", "默认 101"],
        ["飞行开始时间", 0, "s", ""],
        ["飞行结束时间", 300, "s", ""],
        ["飞行时间步长", 5, "s", ""],
        ["可抛质量抛离时间", 120, "s", "默认示例"],
    ]
    for r, row in enumerate(data, 4):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    events = [["子级1", 0, 120, 125, ""], ["子级2", 125, 240, 245, ""], ["子级3", 245, 300, 300, ""], ["助推器", 0, 90, 95, "可选"]]
    for r, row in enumerate(events, 4):
        for c, v in enumerate(row, 6):
            ws.cell(r, c, v)
    mark_inputs(ws, 4, 10, [2, 7, 8, 9])
    style_sheet(ws, {"A": 22, "B": 14, "C": 10, "D": 38, "F": 16, "G": 14, "H": 14, "I": 14, "J": 28})

    ws = wb.create_sheet("输入_不变质量")
    title(ws, "输入_不变质量", 11)
    headers = ["类型", "所属编号", "组件名称", "质量 M kg", "X m", "Y m", "Z m", "Jx kg.m2", "Jy kg.m2", "Jz kg.m2", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    examples = [["子级", 1, "一级结构示例", 10000, 35, 0, 0, 5000, 1200000, 1200000, "示例，可删除"], ["子级", 2, "二级结构示例", 3000, 20, 0, 0, 1200, 180000, 180000, "示例，可删除"], ["子级", 3, "三级结构示例", 1000, 10, 0, 0, 400, 30000, 30000, "示例，可删除"]]
    for r, row in enumerate(examples, 4):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    dv = DataValidation(type="list", formula1='"子级,助推器"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("A4:A303")
    mark_inputs(ws, 4, 303, range(1, 12))
    style_sheet(ws, {"A": 12, "B": 12, "C": 24, "D": 14, "E": 12, "F": 12, "G": 12, "H": 14, "I": 14, "J": 14, "K": 22})

    ws = wb.create_sheet("计算_不变质量")
    title(ws, "计算_不变质量", 10)
    headers = ["类型", "编号", "质量 M", "X", "Y", "Z", "Jx", "Jy", "Jz", "说明"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    entities = [["子级", i, f"第 {i} 子级不变质量"] for i in range(1, 7)] + [["助推器", i, f"第 {i} 个助推器不变质量"] for i in range(1, 9)]
    for r, (typ, idx, desc) in enumerate(entities, 4):
        ws.cell(r, 1, typ)
        ws.cell(r, 2, idx)
        set_formula(ws.cell(r, 3), f'=SUMIFS(输入_不变质量!$D$4:$D$303,输入_不变质量!$A$4:$A$303,A{r},输入_不变质量!$B$4:$B$303,B{r})', True)
        for col, in_col in zip(range(4, 7), ["E", "F", "G"]):
            set_formula(ws.cell(r, col), f'=IF(C{r}=0,0,SUMPRODUCT((输入_不变质量!$A$4:$A$303=A{r})*(输入_不变质量!$B$4:$B$303=B{r})*输入_不变质量!$D$4:$D$303*输入_不变质量!${in_col}$4:${in_col}$303)/C{r})', True)
        set_formula(ws.cell(r, 7), f'=SUMPRODUCT((输入_不变质量!$A$4:$A$303=A{r})*(输入_不变质量!$B$4:$B$303=B{r})*(输入_不变质量!$H$4:$H$303+输入_不变质量!$D$4:$D$303*((E{r}-输入_不变质量!$F$4:$F$303)^2+(F{r}-输入_不变质量!$G$4:$G$303)^2)))', True)
        set_formula(ws.cell(r, 8), f'=SUMPRODUCT((输入_不变质量!$A$4:$A$303=A{r})*(输入_不变质量!$B$4:$B$303=B{r})*(输入_不变质量!$I$4:$I$303+输入_不变质量!$D$4:$D$303*((F{r}-输入_不变质量!$G$4:$G$303)^2+(D{r}-输入_不变质量!$E$4:$E$303)^2)))', True)
        set_formula(ws.cell(r, 9), f'=SUMPRODUCT((输入_不变质量!$A$4:$A$303=A{r})*(输入_不变质量!$B$4:$B$303=B{r})*(输入_不变质量!$J$4:$J$303+输入_不变质量!$D$4:$D$303*((D{r}-输入_不变质量!$E$4:$E$303)^2+(E{r}-输入_不变质量!$F$4:$F$303)^2)))', True)
        ws.cell(r, 10, desc)
    style_sheet(ws, {"A": 12, "B": 10, "C": 14, "D": 12, "E": 12, "F": 12, "G": 14, "H": 14, "I": 14, "J": 30})

    ws = wb.create_sheet("输入_贮箱参数")
    title(ws, "输入_贮箱参数", 22)
    headers = ["类型", "所属编号", "贮箱编号", "名称", "密度rho_phi", "气体rho_g", "秒流量", "加注量Mf", "关机剩余Mr", "Lx/Loki", "Ly", "Lz", "点火tf", "关机tb", "分离ts", "总高H", "总容积Vo", "满箱Xo", "满箱Jyo", "Y偏置", "Z偏置", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    rows = [["子级", 1, 1, "一级氧化剂箱", 1140, 5, 50, 6000, 300, 45, 0, 0, 0, 120, 125, 12, 5.3, 6, 90000, 0, 0, "示例"], ["子级", 1, 2, "一级燃料箱", 810, 4, 35, 3500, 200, 32, 0, 0, 0, 120, 125, 10, 4.3, 5, 50000, 0, 0, "示例"]]
    for r, row in enumerate(rows, 4):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    dv = DataValidation(type="list", formula1='"子级,助推器"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("A4:A103")
    mark_inputs(ws, 4, 103, range(1, 23))
    style_sheet(ws, {get_column_letter(i): 13 for i in range(1, 23)})

    ws = wb.create_sheet("计算_液位积分")
    title(ws, "计算_液位积分", 17)
    headers = ["液位点", "h/H", "h m", "截面积A m2", "体积V_phi", "质量M_phi", "X_phi", "Jy_phi", "Jz_phi", "气体Vg", "气体Mg", "气体Xg", "合成Mv", "合成Xv", "合成Yv", "合成Zv", "说明"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    for r in range(4, 105):
        i = r - 4
        ws.cell(r, 1, i)
        set_formula(ws.cell(r, 2), f'=A{r}/100')
        set_formula(ws.cell(r, 3), f'=B{r}*输入_贮箱参数!$P$4', True)
        set_formula(ws.cell(r, 4), f'=IF(OR(B{r}<0,B{r}>1),0,PI()*输入_贮箱参数!$Q$4/输入_贮箱参数!$P$4)', True)
        if r == 4:
            set_formula(ws.cell(r, 5), "=0")
            set_formula(ws.cell(r, 7), "=0")
            set_formula(ws.cell(r, 8), "=0")
        else:
            set_formula(ws.cell(r, 5), f'=E{r-1}+0.5*(D{r-1}+D{r})*(C{r}-C{r-1})')
            set_formula(ws.cell(r, 7), f'=IF(E{r}=0,0,(SUMPRODUCT($D$4:D{r},$C$4:C{r})/SUM($D$4:D{r})))')
            set_formula(ws.cell(r, 8), f'=F{r}*(C{r}^2+输入_贮箱参数!$Q$4/PI())/12')
        set_formula(ws.cell(r, 6), f'=E{r}*输入_贮箱参数!$E$4', True)
        set_formula(ws.cell(r, 9), f'=H{r}')
        set_formula(ws.cell(r, 10), f'=MAX(0,输入_贮箱参数!$Q$4-E{r})', True)
        set_formula(ws.cell(r, 11), f'=J{r}*输入_贮箱参数!$F$4', True)
        set_formula(ws.cell(r, 12), f'=IF(J{r}=0,0,(输入_贮箱参数!$Q$4*输入_贮箱参数!$R$4-E{r}*G{r})/J{r})', True)
        set_formula(ws.cell(r, 13), f'=F{r}+K{r}')
        set_formula(ws.cell(r, 14), f'=IF(M{r}=0,0,(F{r}*G{r}+K{r}*L{r})/M{r})')
        set_formula(ws.cell(r, 15), '=输入_贮箱参数!$T$4', True)
        set_formula(ws.cell(r, 16), '=输入_贮箱参数!$U$4', True)
        ws.cell(r, 17, "示例按第1个贮箱计算；复制区域可扩展多贮箱")
    style_sheet(ws, {get_column_letter(i): 14 for i in range(1, 18)})

    ws = wb.create_sheet("输入_可抛质量")
    title(ws, "输入_可抛质量", 10)
    headers = ["名称", "质量 M", "X", "Y", "Z", "Jx", "Jy", "Jz", "抛离时间", "备注"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    ws.append(["整流罩示例", 500, 8, 0, 0, 200, 5000, 5000, 120, "示例，可删除"])
    mark_inputs(ws, 4, 103, range(1, 11))
    style_sheet(ws, {"A": 20, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 12, "I": 14, "J": 22})

    ws = wb.create_sheet("计算_时间序列")
    title(ws, "计算_时间序列", 20)
    headers = ["t", "Mu", "Xu", "Yu", "Zu", "Jxu", "Jyu", "Jzu", "Mthw", "Xthw", "Ythw", "Zthw", "Jxthw", "Jythw", "Jzthw", "Mv示例", "Xv示例", "Yv示例", "Zv示例", "说明"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    for r in range(4, 105):
        if r == 4:
            set_formula(ws.cell(r, 1), '=输入_总体参数!$B$7', True)
        else:
            set_formula(ws.cell(r, 1), f'=IF(A{r-1}+输入_总体参数!$B$9>输入_总体参数!$B$8,"",A{r-1}+输入_总体参数!$B$9)', True)
        set_formula(ws.cell(r, 2), '=SUM(计算_不变质量!$C$4:$C$9)', True)
        set_formula(ws.cell(r, 3), f'=IF(B{r}=0,0,SUMPRODUCT(计算_不变质量!$C$4:$C$9,计算_不变质量!$D$4:$D$9)/B{r})', True)
        set_formula(ws.cell(r, 4), f'=IF(B{r}=0,0,SUMPRODUCT(计算_不变质量!$C$4:$C$9,计算_不变质量!$E$4:$E$9)/B{r})', True)
        set_formula(ws.cell(r, 5), f'=IF(B{r}=0,0,SUMPRODUCT(计算_不变质量!$C$4:$C$9,计算_不变质量!$F$4:$F$9)/B{r})', True)
        set_formula(ws.cell(r, 6), '=SUM(计算_不变质量!$G$4:$G$9)', True)
        set_formula(ws.cell(r, 7), '=SUM(计算_不变质量!$H$4:$H$9)', True)
        set_formula(ws.cell(r, 8), '=SUM(计算_不变质量!$I$4:$I$9)', True)
        set_formula(ws.cell(r, 9), f'=SUMIFS(输入_可抛质量!$B$4:$B$103,输入_可抛质量!$I$4:$I$103,">="&A{r})', True)
        for c, src in zip(range(10, 15), ["C", "D", "E", "F", "G"]):
            set_formula(ws.cell(r, c), f'=IF(I{r}=0,0,SUMPRODUCT((输入_可抛质量!$I$4:$I$103>=A{r})*输入_可抛质量!$B$4:$B$103*输入_可抛质量!${src}$4:${src}$103)/I{r})', True)
        set_formula(ws.cell(r, 15), f'=IF(I{r}=0,0,SUMPRODUCT((输入_可抛质量!$I$4:$I$103>=A{r})*输入_可抛质量!$B$4:$B$103*输入_可抛质量!$H$4:$H$103)/I{r})', True)
        set_formula(ws.cell(r, 16), f'=IF(A{r}="",0,MAX(0,输入_贮箱参数!$H$4-输入_贮箱参数!$G$4*MAX(0,A{r}-输入_贮箱参数!$M$4)))', True)
        set_formula(ws.cell(r, 17), '=输入_贮箱参数!$J$4-输入_贮箱参数!$R$4', True)
        set_formula(ws.cell(r, 18), '=输入_贮箱参数!$T$4', True)
        set_formula(ws.cell(r, 19), '=输入_贮箱参数!$U$4', True)
        ws.cell(r, 20, "时间序列页含简化示例；多贮箱可按同结构扩展")
    style_sheet(ws, {get_column_letter(i): 13 for i in range(1, 21)})

    ws = wb.create_sheet("输出_全箭结果")
    title(ws, "输出_全箭结果", 8)
    headers = ["t s", "M kg", "X m", "Y m", "Z m", "Jx kg.m2", "Jy kg.m2", "Jz kg.m2"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    for r in range(4, 105):
        set_formula(ws.cell(r, 1), f'=计算_时间序列!A{r}', True)
        set_formula(ws.cell(r, 2), f'=SUM(计算_时间序列!B{r},计算_时间序列!I{r},计算_时间序列!P{r})', True)
        set_formula(ws.cell(r, 3), f'=IF(B{r}=0,0,(计算_时间序列!B{r}*计算_时间序列!C{r}+计算_时间序列!I{r}*计算_时间序列!J{r}+计算_时间序列!P{r}*计算_时间序列!Q{r})/B{r})', True)
        set_formula(ws.cell(r, 4), f'=IF(B{r}=0,0,(计算_时间序列!B{r}*计算_时间序列!D{r}+计算_时间序列!I{r}*计算_时间序列!K{r}+计算_时间序列!P{r}*计算_时间序列!R{r})/B{r})', True)
        set_formula(ws.cell(r, 5), f'=IF(B{r}=0,0,(计算_时间序列!B{r}*计算_时间序列!E{r}+计算_时间序列!I{r}*计算_时间序列!L{r}+计算_时间序列!P{r}*计算_时间序列!S{r})/B{r})', True)
        set_formula(ws.cell(r, 6), f'=计算_时间序列!F{r}+计算_时间序列!M{r}+计算_时间序列!P{r}*((D{r}-计算_时间序列!R{r})^2+(E{r}-计算_时间序列!S{r})^2)', True)
        set_formula(ws.cell(r, 7), f'=计算_时间序列!G{r}+计算_时间序列!N{r}+计算_时间序列!P{r}*((C{r}-计算_时间序列!Q{r})^2+(E{r}-计算_时间序列!S{r})^2)', True)
        set_formula(ws.cell(r, 8), f'=计算_时间序列!H{r}+计算_时间序列!O{r}+计算_时间序列!P{r}*((C{r}-计算_时间序列!Q{r})^2+(D{r}-计算_时间序列!R{r})^2)', True)
    style_sheet(ws, {"A": 12, "B": 14, "C": 12, "D": 12, "E": 12, "F": 14, "G": 14, "H": 14})

    ws = wb.create_sheet("输出_图表")
    title(ws, "输出_图表", 10)
    ws.cell(3, 1, "图表自动引用 输出_全箭结果")
    for idx, (name, min_col, max_col, pos) in enumerate([("质量-时间", 2, 2, "A5"), ("质心-时间", 3, 5, "A20"), ("转动惯量-时间", 6, 8, "A35")]):
        chart = LineChart()
        chart.title = name
        chart.y_axis.title = name
        chart.x_axis.title = "t/s"
        data = Reference(wb["输出_全箭结果"], min_col=min_col, max_col=max_col, min_row=3, max_row=104)
        cats = Reference(wb["输出_全箭结果"], min_col=1, min_row=4, max_row=104)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, pos)
    style_sheet(ws, {get_column_letter(i): 14 for i in range(1, 11)})

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, (int, float)):
                    c.number_format = '0.000'
                if isinstance(c.value, str) and c.value.startswith('='):
                    c.number_format = '0.000'

    wb.save(OUT)


if __name__ == "__main__":
    build()
    wb = load_workbook(OUT, data_only=False)
    print({"saved": OUT, "sheets": wb.sheetnames})
